from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Alignment, colors, PatternFill, Font, Fill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
import pandas as pd
from filepaths import *
from datetime import date, timedelta
import os
import requests


class InitialiseDb:
    def __init__(self, share_list=False, base_url='', s_p='', stored_path=''):
        """
        Gets the names of the input stocks, initialise formatting and other
        structures.
        """
        date_today = date.today() - timedelta(days=1)
        self.url_date = date_today.strftime("%d%m%Y")
        self.base_url = base_url
        self.input_names = []
        if not share_list:
            self.get_file()
            self.get_info()
        else:
            self.share_path = s_p
            self.stored_shares()
        wb = Workbook()
        self.ws = wb.active
        self.border = Border(left=Side(border_style='thin', color='000000'),
                             right=Side(border_style='thin', color='000000'),
                             top=Side(border_style='thin', color='000000'),
                             bottom=Side(border_style='thin', color='000000'))
        self.ft = Font(color='FFFFFF', bold=True, name='Times New Roman')
        self.allign_style = 'center'
        self.cell_range = 'A1:A{}'.format(2+len(self.input_names))
        self.ws.merge_cells('A1:A2')
        self.ws['A1'] = 'STOCKS'
        self.store_names()
        self.stylise_cells()
        wb.save(stored_path)
        if not share_list:
            os.remove(self.file_path)

    def get_file(self):
        """
        Gets the url for the csv file.
        """
        self.url = self.base_url + 'sec_bhavdata_full_{}.csv'.format(self.url_date)
        values = requests.get(self.url)
        self.file_path = 'sec_bhavdata_full_{}.csv'.format(self.url_date)
        fhand = open(self.file_path, 'wb')
        fhand.write(values.content)
        fhand.close()

    def get_info(self):
        """
        Filters out the shares to only give those with EQ parameter
        """
        raw_input_df = pd.read_csv(self.file_path, sep=r'\s*,\s*', engine='python')
        input_df = raw_input_df[raw_input_df["SERIES"] == 'EQ']
        self.input_names = input_df['SYMBOL']

    def stored_shares(self):
        """
        Gets the names of shares from a share list.
        """
        input_df = pd.read_csv(self.share_path, header=None)
        self.input_names = input_df[0].values.tolist()

    def store_names(self):
        """
        Adds names of initial stocks to the excel sheet.
        """
        st_row, st_col = 3, 'A'
        for name in self.input_names:
            curr_cell = f'{st_col}{st_row}'
            self.ws[curr_cell] = name
            st_row += 1

    def stylise_cells(self):
        """
        Called upon by init after getting  to stylise the rows
        """
        rows = self.ws[self.cell_range]
        for row in rows:
            for cell in row:
                cell.border = self.border
                cell.alignment = Alignment(horizontal=self.allign_style,
                                           vertical=self.allign_style)


def main():
    for i in range(1):
        base_url = data['base_url'][i]
        s_p = data['share_path'][i]
        stored_path = data['stored_path'][i]
        obj = InitialiseDb(True, base_url, s_p, stored_path)


if __name__ == '__main__':
    main()
