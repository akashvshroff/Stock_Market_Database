from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Alignment, colors, PatternFill, Font, Fill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
import pandas as pd
from filepaths import *
import requests
from datetime import date, timedelta, datetime
import os
import sys


class StoreData:
    def __init__(self, stock_list=False, n=1):
        """
        Initialises the various dataframes, stylistic variables and data
        structures for the program! Accepts the parameters that are to be
        stored.
        As reports are released for the prev day then the current day is
        considered to be yesterday.
        """
        flag = False
        date_today = date.today() - timedelta(days=n)
        self.d1 = date_today.strftime("%d-%m-%Y")
        if date_today.weekday() in [5, 6]:
            print('No reports available for Saturday or Sunday.')
            flag = True
        if not flag:
            self.url_date = date_today.strftime("%d%m%Y")
            self.base_url = 'https://archives.nseindia.com/products/content/'
            self.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
            self.ft = Font(color='FFFFFF', bold=True, name='Times New Roman')
            self.allign_style = 'center'
            self.parameters = ['DELIV_PER']
            self.start_row = 1
            self.stored_names, self.input_data, self.input_names,  self.url = [], [], [], '',
            self.cells_ref = ['' for i in range(len(self.parameters))]
            self.share_path = ''
            self.pre_list = False
            if stock_list:
                self.share_path = share_path
                self.pre_list = True
            flag = self.get_file()
            if not flag:  # check if report exists
                wb = load_workbook(stored_path)
                self.ws = wb.active
                self.date_column()
                self.enter_data()
                wb.save(stored_path)
                os.remove(self.file_path)
                print("Report generated for {}".format(self.d1))

    def get_file(self):
        """
        Retrieves the url for the file that is to be scraped.
        """
        self.url = self.base_url + 'sec_bhavdata_full_{}.csv'.format(self.url_date)
        # print(self.url)
        values = requests.get(self.url)
        if values.status_code == 404:
            print("No report available. Try tomorrow. Exiting.")
            return True
        self.file_path = 'sec_bhavdata_full_{}.csv'.format(self.url_date)
        fhand = open(self.file_path, 'wb')
        fhand.write(values.content)
        fhand.close()
        self.retrieve_data()

    def retrieve_data(self):
        """
        Scrapes the data retrieved, adds it to the data frames and parses the
        data.
        """
        raw_input_df = pd.read_csv(self.file_path, sep=r'\s*,\s*', engine='python')
        input_df = raw_input_df[raw_input_df["SERIES"] == 'EQ']
        self.input_names = input_df['SYMBOL'].values.tolist()
        self.input_data = input_df[self.parameters]
        self.input_data.reset_index(inplace=True, drop=True)
        if not self.pre_list:
            stored_data = pd.read_excel(stored_path)
            self.stored_names = stored_data['STOCKS']
        else:
            input_df = pd.read_csv(self.share_path, header=None)
            self.stored_names = input_df[0].values.tolist()

    def date_column(self):
        """
        Calculates which column is to be used for the day. Fills in the date
        that is to be used for the present day. Finds the next columns as well,
        depending on the lenght of the parameters.
        """
        rows = self.ws.iter_rows(min_row=self.start_row,
                                 max_row=self.start_row)  # the row where your headings are
        row = next(rows)
        headings = [c.value for c in row]
        col_letter = ''
        for col, heading in enumerate(headings):
            if heading == self.d1 or heading is None:
                for i in range(len(self.parameters)):
                    self.cells_ref[i] = get_column_letter(col+i+1)
        if not all(self.cells_ref):
            for i in range(len(self.parameters)):
                self.cells_ref[i] = get_column_letter(len(headings)+i+1)
        self.enter_initial()

    def enter_initial(self):
        """
        Enters the date, as well as the parameters list into the sheet.
        """
        start_cell, end_cell = f'{self.cells_ref[0]}{self.start_row}', f'{self.cells_ref[-1]}{self.start_row}'
        merge = '{}:{}'.format(start_cell, end_cell)
        self.ws.merge_cells(merge)
        self.ws[start_cell] = self.d1
        self.ws[start_cell].border = self.border
        self.ws[start_cell].alignment = Alignment(
            horizontal=self.allign_style, vertical=self.allign_style)
        curr_row = self.start_row + 1
        for i, para in enumerate(self.parameters):
            curr_cell = f'{self.cells_ref[i]}{curr_row}'
            self.ws[curr_cell] = para

    def stylise_cells(self, cell_range):
        """
        Stylises cells within a range with the border, allignment etc.
        """
        rows = self.ws[cell_range]
        for row in rows:
            for cell in row:
                cell.border = self.border
                cell.alignment = Alignment(horizontal=self.allign_style,
                                           vertical=self.allign_style)

    def enter_data(self):
        """
        Enters the data in the excel sheet based on the names that are already
        stored and those that are new are apended to the end of the STOCKS list
        on the first column.
        """
        num_stored = len(self.stored_names)
        curr_row = self.start_row + 2
        rem_names = self.input_names[::]  # to remove names
        for stock in self.stored_names:
            if stock in self.input_names:  # there is data for it
                stock_index = self.input_names.index(stock)
                for num, parameter in enumerate(self.parameters):
                    curr_cell = '{}{}'.format(self.cells_ref[num], curr_row)
                    value = self.input_data.at[stock_index, parameter]
                    self.ws[curr_cell] = value
                rem_names.remove(stock)
                self.input_data.drop([stock_index], inplace=True)
            else:  # missing from retrieved list.
                for num, parameter in enumerate(self.parameters):
                    curr_cell = '{}{}'.format(self.cells_ref[num], curr_row)
                    self.ws[curr_cell] = '-'
            curr_row += 1
        cell_range = '{}{}:{}{}'.format(
            self.cells_ref[0], self.start_row+1, self.cells_ref[-1], self.start_row+num_stored+1)
        self.stylise_cells(cell_range)
        if not self.pre_list:  # dealing with all the shares
            if rem_names:
                # now it is only the cells that haven't been added before
                start_row = 2 + num_stored
                curr_row = start_row
                for name in rem_names:
                    curr_cell = f'A{curr_row}'
                    self.ws[curr_cell] = name
                    stock_index = self.input_names.index(name)
                    for num, parameter in enumerate(self.parameters):
                        curr_cell = '{}{}'.format(self.cells_ref[num], curr_row)
                        value = self.input_data.at[stock_index, parameter]
                        self.ws[curr_cell] = value
                    curr_row += 1
            cell_ranges = [f'A{1}:A{curr_row-1}', f'B{2}:{self.cells_ref[-1]}{curr_row-1}']
            for cell_range in cell_ranges:
                self.stylise_cells(cell_range)


def main():
    StoreData(True, 1)


if __name__ == '__main__':
    main()
